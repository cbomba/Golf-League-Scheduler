from flask import Flask, render_template, request, jsonify, send_file
from scheduler import generate_schedule, summarize_schedule
from datetime import datetime, date, timedelta
import openpyxl
from io import BytesIO

app = Flask(__name__)


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate():

    data = request.json

    men_enabled = data.get("menEnabled", False)
    women_enabled = data.get("womenEnabled", False)

    men_count = int(data.get("menCount") or 0)
    women_count = int(data.get("womenCount") or 0)

    start_time = data.get("startTime")

    league_start_str = data.get("leagueStartDate")
    league_end_str = data.get("leagueEndDate")

    placement = data.get("placement", "mixed_alternating")

    #
    # validation
    #

    if not start_time:
        return jsonify({"error": "Please choose a first tee time."}), 400

    if not league_start_str or not league_end_str:
        return jsonify({"error": "Please choose league start and end dates."}), 400

    if not men_enabled and not women_enabled:
        return jsonify({"error": "Please select at least one league."}), 400

    #
    # convert dates
    #

    league_start = datetime.strptime(league_start_str, "%Y-%m-%d").date()
    league_end = datetime.strptime(league_end_str, "%Y-%m-%d").date()
    days_of_week = data.get("daysOfWeek", [])
    
    weekday_map = {
        "mon":0,
        "tue":1,
        "wed":2,
        "thu":3,
        "fri":4,
        "sat":5,
        "sun":6
    }

    weekdays = [weekday_map[d] for d in days_of_week]

    #
    # build player IDs
    #

    men = [str(i+1) for i in range(men_count)] if men_enabled else []
    women = [chr(65+i) for i in range(women_count)] if women_enabled else []

    try:

        season = generate_schedule(
            men_players=men,
            women_players=women,
            start_date=league_start,
            end_date=league_end,
            first_tee_time=start_time,
            weekdays=weekdays,
            placement_mode=placement
        )

        summary = summarize_schedule(season, men, women)

        return jsonify({
            "season":[week.__dict__ for week in season],
            "summary":summary
        })

    except Exception as e:
        return jsonify({"error":str(e)}),400


@app.route("/export", methods=["POST"])
def export_excel():

    data = request.json
    season = data["season"]
    player_map = data["playerMap"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    row = 1

    #
    # HEADER
    #

    ws.cell(row=row, column=1).value = "Date"

    col = 2
    for slot in season[0]["slots"]:
        ws.cell(row=row, column=col).value = slot["time"]
        ws.cell(row=row, column=col + 1).value = slot["time"]
        col += 2

    row += 1

    #
    # SCHEDULE
    #

    for week in season:

        ws.cell(row=row, column=1).value = week["play_date"]

        col = 2

        for slot in week["slots"]:

            group = slot["group"]

            if len(group) == 4:
                ws.cell(row=row, column=col).value = f"{group[0]}-{group[1]}"
                ws.cell(row=row, column=col + 1).value = f"{group[2]}-{group[3]}"

            elif len(group) == 3:
                ws.cell(row=row, column=col).value = f"{group[0]}-{group[1]}"
                ws.cell(row=row, column=col + 1).value = group[2]

            elif len(group) == 2:
                ws.cell(row=row, column=col).value = group[0]
                ws.cell(row=row, column=col + 1).value = group[1]

            else:
                ws.cell(row=row, column=col).value = "-".join(group)

            col += 2

        row += 1

    #
    # PLAYER LIST
    #

    row += 2
    ws.cell(row=row, column=1).value = "Player List"

    row += 1

    for key in sorted(player_map.keys(), key=lambda x: (not x.isdigit(), x)):
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = player_map[key]
        row += 1

    #
    # SAVE FILE
    #

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="golf_league_schedule.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(debug=True)